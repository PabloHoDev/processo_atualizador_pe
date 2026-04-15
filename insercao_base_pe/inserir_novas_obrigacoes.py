import pandas as pd
from openpyxl import load_workbook

from configuracoes.parametros_processo import (
    CAMINHO_BASE_GERAL_PE,
    COLUNA_CHAVE
)


# ===============================
# 📌 DEFINIR ABA DE DESTINO
# ===============================

def definir_aba_destino(linha: pd.Series) -> str:
    """
    Define em qual aba a obrigação será inserida.
    Regra atual: baseada na UF PRESTADOR
    """
    uf = linha.get("UF PRESTADOR", "").strip()

    if not uf:
        raise ValueError("UF PRESTADOR não informada")

    return uf


# ===============================
# 📍 ENCONTRAR PRÓXIMA LINHA
# ===============================

def proxima_linha_vazia(ws) -> int:
    """
    Retorna a próxima linha vazia da aba.
    """
    return ws.max_row + 1


# ===============================
# 🚀 INSERÇÃO PRINCIPAL
# ===============================

def inserir_novas_obrigacoes(df_novos: pd.DataFrame):
    """
    Insere novas obrigações na Base Geral PE.
    """

    if df_novos.empty:
        print("Nenhuma nova obrigação para inserir.")
        return

    wb = load_workbook(CAMINHO_BASE_GERAL_PE)

    inseridos = 0
    erros = []

    for _, linha in df_novos.iterrows():
        try:
            aba_nome = definir_aba_destino(linha)

            if aba_nome not in wb.sheetnames:
                raise ValueError(f"Aba '{aba_nome}' não existe na base")

            ws = wb[aba_nome]

            linha_excel = proxima_linha_vazia(ws)

            # ===============================
            # 📥 INSERIR DADOS
            # ===============================
            for col_idx, col_name in enumerate(df_novos.columns, start=1):
                ws.cell(row=linha_excel, column=col_idx, value=linha[col_name])

            inseridos += 1

        except Exception as e:
            erros.append({
                "obrigacao": linha.get(COLUNA_CHAVE),
                "erro": str(e)
            })

    # ===============================
    # 💾 SALVAR ARQUIVO
    # ===============================
    wb.save(CAMINHO_BASE_GERAL_PE)

    return {
        "inseridos": inseridos,
        "erros": erros
    }